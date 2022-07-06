import openpyxl

# 경로 저장
f_path = r'C:\Users\shin7\OneDrive\바탕 화면\Crawling\엑셀 다루기\강아지_data.xlsx'

# 기존 파일 읽기
wb = openpyxl.load_workbook(f_path)

# 파일 시트 가져오기
ws = wb['Test1']


ws['A3'] = '2'
ws['B3'] = '달래'

wb.save(f_path)