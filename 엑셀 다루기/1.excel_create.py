import openpyxl

# 1)엑셀 만들기
wb = openpyxl.Workbook()

# 2)엑셀 워크시트 만들기
ws = wb.create_sheet('Test1')

# 3)데이터 추가
ws['A1'] = '순서'
ws['B1'] = '이름'

ws['A2'] = '1'
ws['B2'] = '봉남'

# 4)엑셀 저장
wb.save(r'C:\Users\shin7\OneDrive\바탕 화면\Crawling\엑셀 다루기\강아지_data.xlsx')
